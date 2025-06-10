"""
DICOMMamographyExtractor.py - Extração direta de DICOM SR de Mamografia para Excel

Este script navega recursivamente por estruturas de pastas, encontra arquivos DICOM SR
de dose de radiação de mamografia e gera diretamente a planilha Excel sem JSON intermediário.
Extrai apenas os campos necessários para otimizar performance.
MODIFICADO: Salva valores numéricos como números puros (sem unidades) no Excel.
"""

import pydicom
import os
import argparse
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class DICOMMamographyExtractor:
    """Extrator direto de DICOM SR de Mamografia para Excel"""

    def __init__(self):
        # Códigos DICOM específicos para mamografia
        self.concept_codes = {
            # Dados de irradiação acumulados
            'accumulated_dose_data': '113702',
            'accumulated_agd': '111637',  # Accumulated Average Glandular Dose

            # Eventos de irradiação
            'irradiation_event': '113706',
            'irradiation_event_uid': '113769',
            'datetime_started': '111526',
            'irradiation_event_type': '113721',
            'acquisition_protocol': '125203',

            # Anatomia e lateralidade
            'anatomical_structure': 'T-D0005',
            'laterality': 'G-C171',
            'image_view': '111031',
            'target_region': '123014',

            # Parâmetros técnicos
            'kvp': '113733',
            'tube_current': '113734',
            'exposure_time': '113824',
            'pulse_width': '113793',
            'number_of_pulses': '113768',
            'irradiation_duration': '113742',
            'focal_spot_size': '113766',

            # Dose e exposição
            'average_glandular_dose': '111631',  # AGD por evento
            'entrance_exposure': '111636',
            'half_value_layer': '111634',
            'reference_point_def': '113780',

            # Geometria
            'compression_thickness': '111633',
            'distance_source_to_rp': '113737',
            'collimated_field_area': '113790',
            'collimated_field_height': '113788',
            'collimated_field_width': '113789',

            # Equipamento
            'anode_target_material': '111632',
            'xray_filters': '113771',
            'filter_type': '113772',
            'filter_material': '113757',
            'xray_grid': '111635',
            'mechanical_config': '113956',
            'positioner_angle': '112011',

            # Aquisição
            'acquisition_plane': '113764',
            'acquired_image': '113795',
            'dose_source': '113854'  # Source of Dose Information
        }

    def find_dicom_files_recursive(self, root_path: str, debug_mode: bool = False) -> list:
        """Busca recursivamente por arquivos DICOM SR"""
        dicom_files = []

        if debug_mode:
            print(f"🔍 Buscando arquivos DICOM em: {root_path}")

        try:
            for root, dirs, files in os.walk(root_path):
                if root == root_path:
                    continue

                for file in files:
                    file_path = os.path.join(root, file)

                    if self.is_dicom_sr_file(file_path):
                        dicom_files.append(file_path)
                        if debug_mode:
                            print(f"  ✓ DICOM encontrado: {file_path}")

        except Exception as e:
            if debug_mode:
                print(f"❌ Erro na busca: {str(e)}")

        return dicom_files

    def is_dicom_sr_file(self, file_path: str) -> bool:
        """Verifica se é um DICOM SR válido rapidamente"""
        try:
            if not os.path.isfile(file_path) or os.path.getsize(file_path) < 132:
                return False

            # Verifica prefixo DICM
            with open(file_path, 'rb') as f:
                f.seek(128)
                if f.read(4) != b'DICM':
                    return False

            # Leitura mínima para verificar se é SR de mamografia
            ds = pydicom.dcmread(file_path, stop_before_pixels=True, force=True)

            # Verifica se é SR e se contém dados de mamografia
            if not (hasattr(ds, 'Modality') and ds.Modality == 'SR' and
                    hasattr(ds, 'ContentSequence')):
                return False

            # Verifica se contém dados de dose de mamografia
            return self.contains_mammography_data(ds)

        except:
            return False

    def contains_mammography_data(self, ds) -> bool:
        """Verifica se o DICOM contém dados específicos de mamografia"""
        try:
            if not hasattr(ds, 'ContentSequence'):
                return False

            # Procura por códigos específicos de mamografia
            for item in ds.ContentSequence:
                if (hasattr(item, 'ConceptNameCodeSequence') and
                        item.ConceptNameCodeSequence):
                    code = getattr(item.ConceptNameCodeSequence[0], 'CodeValue', '')

                    # Verifica se contém dados de dose acumulada ou eventos de mamografia
                    if code in [self.concept_codes['accumulated_dose_data'],
                                self.concept_codes['irradiation_event']]:
                        return True

                    # Verifica se é um relatório de dose de raios-X
                    if (code == '113701' and hasattr(item, 'ContentSequence')):
                        for sub_item in item.ContentSequence:
                            if hasattr(sub_item, 'ConceptCodeSequence') and sub_item.ConceptCodeSequence:
                                meaning = getattr(sub_item.ConceptCodeSequence[0], 'CodeMeaning', '')
                                if 'Mammography' in meaning:
                                    return True
            return False
        except:
            return False

    def calculate_age(self, birth_date_str: str, exam_date_str: str):
        """Calcula idade do paciente"""
        if not birth_date_str or not exam_date_str:
            return '-'

        try:
            birth_date = None
            exam_date = None

            # Formatos de data
            date_formats = ['%b %d, %Y', '%B %d, %Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']
            exam_formats = date_formats + ['%b %d, %Y, %I:%M:%S %p', '%B %d, %Y, %I:%M:%S %p']

            # Parse data nascimento
            for fmt in date_formats:
                try:
                    birth_date = datetime.strptime(birth_date_str.strip(), fmt)
                    break
                except ValueError:
                    continue

            # Parse data exame
            for fmt in exam_formats:
                try:
                    exam_date = datetime.strptime(exam_date_str.strip(), fmt)
                    break
                except ValueError:
                    continue

            # Fallback: extrai apenas anos
            if not birth_date:
                birth_match = re.search(r'(\d{4})', birth_date_str)
                if birth_match:
                    birth_date = datetime(int(birth_match.group(1)), 1, 1)

            if not exam_date:
                exam_match = re.search(r'(\d{4})', exam_date_str)
                if exam_match:
                    exam_date = datetime(int(exam_match.group(1)), 6, 15)

            if birth_date and exam_date:
                age = exam_date.year - birth_date.year
                if (exam_date.month, exam_date.day) < (birth_date.month, birth_date.day):
                    age -= 1
                return age

        except:
            pass

        return '-'

    def format_date(self, date_str: str) -> str:
        """Formata data DICOM"""
        if not date_str or len(date_str) < 8:
            return ""

        try:
            year = date_str[:4]
            month = date_str[4:6]
            day = date_str[6:8]

            months = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            month_name = months[int(month)]
            return f"{month_name} {int(day)}, {year}"
        except:
            return date_str

    def find_content_by_code(self, content_sequence, code_value: str):
        """Encontra item por código DICOM"""
        for item in content_sequence:
            try:
                if (hasattr(item, 'ConceptNameCodeSequence') and
                        item.ConceptNameCodeSequence):
                    concept_code = getattr(item.ConceptNameCodeSequence[0], 'CodeValue', '')
                    if concept_code == code_value:
                        return item
            except:
                continue
        return None

    def get_text_value(self, content_item) -> str:
        """Extrai valor de texto"""
        return getattr(content_item, 'TextValue', '')

    def get_code_meaning(self, content_item) -> str:
        """Extrai code meaning"""
        try:
            if hasattr(content_item, 'ConceptCodeSequence') and content_item.ConceptCodeSequence:
                return getattr(content_item.ConceptCodeSequence[0], 'CodeMeaning', '')
        except:
            pass
        return ""

    def get_numeric_value_with_unit(self, content_item) -> str:
        """Extrai valor numérico com unidade"""
        try:
            if hasattr(content_item, 'MeasuredValueSequence') and content_item.MeasuredValueSequence:
                measured_value = content_item.MeasuredValueSequence[0]
                numeric_value = getattr(measured_value, 'NumericValue', '')

                unit = ''
                if (hasattr(measured_value, 'MeasurementUnitsCodeSequence') and
                        measured_value.MeasurementUnitsCodeSequence):
                    unit_seq = measured_value.MeasurementUnitsCodeSequence[0]
                    unit = getattr(unit_seq, 'CodeMeaning', '')

                if numeric_value and unit:
                    return f"{numeric_value} {unit}"
                elif numeric_value:
                    return str(numeric_value)
        except:
            pass
        return ""

    def get_numeric_value_only(self, content_item) -> str:
        """Extrai apenas o valor numérico sem unidade"""
        try:
            if hasattr(content_item, 'MeasuredValueSequence') and content_item.MeasuredValueSequence:
                measured_value = content_item.MeasuredValueSequence[0]
                numeric_value = getattr(measured_value, 'NumericValue', '')
                return str(numeric_value) if numeric_value else ""
        except:
            pass
        return ""

    def get_numeric_value_as_float(self, content_item):
        """Extrai valor numérico como float para Excel, retorna None se não for número"""
        try:
            if hasattr(content_item, 'MeasuredValueSequence') and content_item.MeasuredValueSequence:
                measured_value = content_item.MeasuredValueSequence[0]
                numeric_value = getattr(measured_value, 'NumericValue', '')
                if numeric_value:
                    return float(numeric_value)
        except:
            pass
        return None

    def safe_numeric_value(self, content_item, return_as_number=False):
        """
        Extrai valor numérico de forma segura
        Args:
            content_item: Item do DICOM
            return_as_number: Se True, retorna float/int para Excel, senão string com unidade
        """
        if return_as_number:
            value = self.get_numeric_value_as_float(content_item)
            return value if value is not None else '-'
        else:
            return self.get_numeric_value_with_unit(content_item) or '-'

    def extract_laterality(self, content_sequence) -> str:
        """Extrai lateralidade (Left/Right)"""
        for item in content_sequence:
            try:
                if (hasattr(item, 'ConceptNameCodeSequence') and
                        item.ConceptNameCodeSequence and
                        getattr(item.ConceptNameCodeSequence[0], 'CodeValue', '') == self.concept_codes['laterality']):

                    if hasattr(item, 'ConceptCodeSequence') and item.ConceptCodeSequence:
                        code_meaning = getattr(item.ConceptCodeSequence[0], 'CodeMeaning', '')
                        if 'Left' in code_meaning:
                            return 'Left'
                        elif 'Right' in code_meaning:
                            return 'Right'
            except:
                continue
        return ""

    def aggregate_multiple_values(self, content_sequence, code_value: str) -> dict:
        """Agrega múltiplos valores do mesmo parâmetro"""
        values = []
        for item in content_sequence:
            try:
                if (hasattr(item, 'ConceptNameCodeSequence') and
                    item.ConceptNameCodeSequence and
                    getattr(item.ConceptNameCodeSequence[0], 'CodeValue', '') == code_value):
                    value = self.get_numeric_value_only(item)
                    if value and value != '':
                        try:
                            values.append(float(value))
                        except:
                            pass
            except:
                continue

        if values:
            return {
                'min': round(min(values), 3),
                'max': round(max(values), 3),
                'avg': round(sum(values)/len(values), 3),
                'count': len(values)
            }
        return {'min': None, 'max': None, 'avg': None, 'count': 0}

    def extract_all_filters(self, content_sequence) -> list:
        """Extrai todos os filtros do evento"""
        filters = []
        for item in content_sequence:
            try:
                if (hasattr(item, 'ConceptNameCodeSequence') and
                    item.ConceptNameCodeSequence and
                    getattr(item.ConceptNameCodeSequence[0], 'CodeValue', '') == self.concept_codes['xray_filters']):

                    if hasattr(item, 'ContentSequence'):
                        for filter_item in item.ContentSequence:
                            if (hasattr(filter_item, 'ConceptNameCodeSequence') and
                                filter_item.ConceptNameCodeSequence and
                                getattr(filter_item.ConceptNameCodeSequence[0], 'CodeValue', '') ==
                                self.concept_codes['filter_material']):
                                material = self.get_code_meaning(filter_item)
                                if material and material not in filters:
                                    filters.append(material)
            except:
                continue
        return filters

    def extract_excel_data(self, dicom_path: str) -> list:
        """Extrai dados específicos de mamografia para o Excel"""
        try:
            ds = pydicom.dcmread(dicom_path)

            if (not hasattr(ds, 'Modality') or ds.Modality != 'SR' or
                    not hasattr(ds, 'ContentSequence')):
                return []

            # Dados básicos do paciente
            patient_id = str(getattr(ds, 'PatientID', ''))
            patient_name = str(getattr(ds, 'PatientName', '')).replace('^', ' ').strip()
            sex = str(getattr(ds, 'PatientSex', ''))

            # Datas
            birth_date_raw = str(getattr(ds, 'PatientBirthDate', ''))
            birth_date = self.format_date(birth_date_raw) if birth_date_raw else ''

            study_date_raw = str(getattr(ds, 'StudyDate', ''))
            study_time_raw = str(getattr(ds, 'StudyTime', ''))
            study_date = self.format_date(study_date_raw) if study_date_raw else ''
            if study_date and study_time_raw and len(study_time_raw) >= 6:
                hour = study_time_raw[:2]
                minute = study_time_raw[2:4]
                second = study_time_raw[4:6]
                study_date = f"{study_date}, {hour}:{minute}:{second}"

            # Calcula idade
            age = self.calculate_age(birth_date, study_date)
            age_value = int(age) if isinstance(age, int) or (isinstance(age, str) and age.isdigit()) else age

            # Patient ID como número se possível
            patient_id_value = int(patient_id) if patient_id and patient_id.isdigit() else (
                patient_id if patient_id else '-')

            # Dados do equipamento
            manufacturer = str(getattr(ds, 'Manufacturer', ''))
            model = str(getattr(ds, 'ManufacturerModelName', ''))
            station_name = str(getattr(ds, 'StationName', ''))

            # Processa conteúdo principal
            excel_rows = []
            main_content = ds.ContentSequence

            # Extrai fonte da informação de dose
            dose_source = ''
            source_item = self.find_content_by_code(main_content, self.concept_codes['dose_source'])
            if source_item:
                dose_source = self.get_code_meaning(source_item)

            # Dicionário para armazenar AGD acumulada por lateralidade
            accumulated_agd = {'Left': None, 'Right': None}

            # Primeiro, extrai dose acumulada por lateralidade
            for item in main_content:
                try:
                    if (hasattr(item, 'ConceptNameCodeSequence') and
                            item.ConceptNameCodeSequence and
                            getattr(item.ConceptNameCodeSequence[0], 'CodeValue', '') ==
                            self.concept_codes['accumulated_dose_data']):

                        if hasattr(item, 'ContentSequence'):
                            for sub_item in item.ContentSequence:
                                try:
                                    if (hasattr(sub_item, 'ConceptNameCodeSequence') and
                                            sub_item.ConceptNameCodeSequence and
                                            getattr(sub_item.ConceptNameCodeSequence[0], 'CodeValue', '') ==
                                            self.concept_codes['accumulated_agd']):

                                        agd_value = self.get_numeric_value_as_float(sub_item)
                                        if agd_value is not None and hasattr(sub_item, 'ContentSequence'):
                                            laterality = self.extract_laterality(sub_item.ContentSequence)
                                            if laterality:
                                                accumulated_agd[laterality] = agd_value
                                except:
                                    continue
                except:
                    continue

            # Processa eventos de irradiação
            events_found = False

            for item in main_content:
                try:
                    if (hasattr(item, 'ConceptNameCodeSequence') and
                            item.ConceptNameCodeSequence and
                            getattr(item.ConceptNameCodeSequence[0], 'CodeValue', '') ==
                            self.concept_codes['irradiation_event']):

                        events_found = True

                        if hasattr(item, 'ContentSequence'):
                            event_content = item.ContentSequence

                            # Dados do evento
                            event_uid = ''
                            datetime_started = ''
                            event_type = ''
                            protocol = ''
                            laterality = ''
                            image_view = ''
                            target_region = ''

                            # Parâmetros técnicos básicos (como números para Excel)
                            kvp = None
                            tube_current = None
                            exposure_time = None
                            pulse_width = None
                            number_of_pulses = None
                            irradiation_duration = None
                            focal_spot_size = None

                            # Análise detalhada dos parâmetros múltiplos
                            kvp_stats = self.aggregate_multiple_values(event_content, self.concept_codes['kvp'])
                            current_stats = self.aggregate_multiple_values(event_content, self.concept_codes['tube_current'])
                            pulse_stats = self.aggregate_multiple_values(event_content, self.concept_codes['pulse_width'])

                            # Dose e exposição (como números para Excel)
                            agd = None
                            entrance_exposure = None
                            half_value_layer = None

                            # Geometria (como números para Excel)
                            compression_thickness = None
                            distance_source_rp = None
                            field_area = None
                            field_height = None
                            field_width = None

                            # Equipamento
                            anode_material = ''
                            grid_type = ''
                            positioner_angle = None

                            # Filtros (múltiplos)
                            all_filters = self.extract_all_filters(event_content)
                            filter_primary = all_filters[0] if len(all_filters) > 0 else ''
                            filter_secondary = all_filters[1] if len(all_filters) > 1 else ''
                            filter_tertiary = all_filters[2] if len(all_filters) > 2 else ''

                            # Extrai dados básicos do evento
                            for event_item in event_content:
                                try:
                                    if not hasattr(event_item,
                                                   'ConceptNameCodeSequence') or not event_item.ConceptNameCodeSequence:
                                        continue

                                    code = getattr(event_item.ConceptNameCodeSequence[0], 'CodeValue', '')

                                    if code == self.concept_codes['irradiation_event_uid']:
                                        event_uid = getattr(event_item, 'UID', '')
                                    elif code == self.concept_codes['datetime_started']:
                                        datetime_started = getattr(event_item, 'DateTime', '')
                                    elif code == self.concept_codes['irradiation_event_type']:
                                        event_type = self.get_code_meaning(event_item)
                                    elif code == self.concept_codes['acquisition_protocol']:
                                        protocol = self.get_text_value(event_item)
                                    elif code == self.concept_codes['anatomical_structure']:
                                        target_region = self.get_code_meaning(event_item)
                                        if hasattr(event_item, 'ContentSequence'):
                                            laterality = self.extract_laterality(event_item.ContentSequence)
                                    elif code == self.concept_codes['image_view']:
                                        image_view = self.get_code_meaning(event_item)
                                    elif code == self.concept_codes['target_region']:
                                        target_region = self.get_code_meaning(event_item)
                                    elif code == self.concept_codes['kvp'] and kvp is None:
                                        kvp = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['tube_current'] and tube_current is None:
                                        tube_current = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['exposure_time']:
                                        exposure_time = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['pulse_width'] and pulse_width is None:
                                        pulse_width = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['number_of_pulses']:
                                        number_of_pulses = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['irradiation_duration']:
                                        irradiation_duration = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['focal_spot_size']:
                                        focal_spot_size = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['average_glandular_dose']:
                                        agd = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['entrance_exposure']:
                                        entrance_exposure = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['half_value_layer']:
                                        half_value_layer = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['compression_thickness']:
                                        compression_thickness = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['distance_source_to_rp']:
                                        distance_source_rp = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['collimated_field_area']:
                                        field_area = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['collimated_field_height']:
                                        field_height = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['collimated_field_width']:
                                        field_width = self.get_numeric_value_as_float(event_item)
                                    elif code == self.concept_codes['anode_target_material']:
                                        anode_material = self.get_code_meaning(event_item)
                                    elif code == self.concept_codes['xray_grid']:
                                        grid_type = self.get_code_meaning(event_item)
                                    elif code == self.concept_codes['positioner_angle']:
                                        positioner_angle = self.get_numeric_value_as_float(event_item)

                                except:
                                    continue

                            # Função para converter None para '-' para campos de texto, manter None para números
                            def safe_text_value(val):
                                return val if val else '-'

                            def safe_numeric_value(val):
                                return val  # None será tratado como célula vazia no Excel

                            # AGD acumulada baseada na lateralidade
                            accumulated_agd_value = accumulated_agd.get(laterality) if laterality else None

                            # Cria linha para Excel
                            excel_row = [
                                patient_id_value,  # ID do paciente
                                patient_name or '-',  # Nome do paciente
                                sex or '-',  # Sexo
                                birth_date or '-',  # Data de nascimento
                                age_value,  # Idade
                                study_date or '-',  # Data do exame
                                manufacturer or '-',  # Fabricante
                                model or '-',  # Modelo do equipamento
                                station_name or '-',  # Nome da estação
                                protocol or '-',  # Protocolo de aquisição
                                laterality or '-',  # Lateralidade
                                image_view or '-',  # Projeção (CC, MLO, etc)
                                event_type or '-',  # Tipo de evento
                                safe_numeric_value(kvp),  # kVp (primeiro valor)
                                kvp_stats['min'],  # kVp mínimo
                                kvp_stats['max'],  # kVp máximo
                                kvp_stats['avg'],  # kVp médio
                                safe_numeric_value(tube_current),  # Corrente do tubo (primeiro valor)
                                current_stats['min'],  # mA mínimo
                                current_stats['max'],  # mA máximo
                                current_stats['avg'],  # mA médio
                                safe_numeric_value(exposure_time),  # Tempo de exposição
                                safe_numeric_value(number_of_pulses),  # Número de pulsos
                                pulse_stats['count'],  # Total de pulsos registrados
                                safe_numeric_value(pulse_width),  # Largura do pulso (primeiro valor)
                                pulse_stats['min'],  # Pulse width mínimo
                                pulse_stats['max'],  # Pulse width máximo
                                pulse_stats['avg'],  # Pulse width médio
                                safe_numeric_value(irradiation_duration),  # Duração da irradiação
                                safe_numeric_value(focal_spot_size),  # Tamanho do ponto focal
                                safe_numeric_value(agd),  # Dose glandular média (evento)
                                accumulated_agd_value,  # Dose glandular acumulada
                                safe_numeric_value(entrance_exposure),  # Exposição na entrada
                                safe_numeric_value(half_value_layer),  # Camada de semi-atenuação
                                safe_numeric_value(compression_thickness),  # Espessura de compressão
                                safe_numeric_value(distance_source_rp),  # Distância fonte-ponto ref
                                safe_numeric_value(field_area),  # Área do campo colimado
                                safe_numeric_value(field_height),  # Altura do campo
                                safe_numeric_value(field_width),  # Largura do campo
                                safe_text_value(anode_material),  # Material do anodo
                                filter_primary or '-',  # Filtro principal
                                filter_secondary or '-',  # Filtro secundário
                                filter_tertiary or '-',  # Filtro terciário
                                safe_text_value(grid_type),  # Tipo de grade
                                safe_numeric_value(positioner_angle),  # Ângulo do posicionador
                                dose_source or '-',  # Fonte da informação de dose
                                event_uid or '-'  # UID do evento
                            ]

                            excel_rows.append(excel_row)

                except:
                    continue

            # Se não encontrou eventos, cria linha básica
            if not events_found:
                excel_row = [
                    patient_id_value, patient_name or '-', sex or '-', birth_date or '-', age_value,
                    study_date or '-', manufacturer or '-', model or '-', station_name or '-',
                    '-', '-', '-', '-', None, None, None, None, None, None, None, None, None, None, None,
                    None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                    '-', '-', '-', '-', '-', None, dose_source or '-', '-'
                ]
                excel_rows.append(excel_row)

            return excel_rows

        except Exception as e:
            return []

    def generate_excel_direct(self, root_path: str, output_file: str, debug_mode: bool = False) -> bool:
        """Gera Excel diretamente dos DICOMs de mamografia"""

        print(f"🔍 Buscando arquivos DICOM de mamografia em: {os.path.abspath(root_path)}")

        # Busca arquivos DICOM
        dicom_files = self.find_dicom_files_recursive(root_path, debug_mode)

        if not dicom_files:
            print("❌ Nenhum arquivo DICOM SR de mamografia encontrado")
            return False

        print(f"📊 Encontrados {len(dicom_files)} arquivos DICOM de mamografia")
        print(f"📄 Gerando Excel: {output_file}")

        # Cria planilha Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatórios DICOM Mamografia"

        # Cabeçalhos específicos para mamografia
        headers = [
            "ID do paciente", "Nome do paciente", "Sexo", "Data de nascimento", "Idade",
            "Data do exame", "Fabricante", "Modelo do equipamento", "Nome da estação",
            "Protocolo de aquisição", "Lateralidade", "Projeção", "Tipo de evento",
            "kVp", "kVp mínimo", "kVp máximo", "kVp médio",
            "Corrente do tubo (mA)", "mA mínimo", "mA máximo", "mA médio",
            "Tempo de exposição", "Número de pulsos", "Total pulsos registrados",
            "Largura do pulso", "Pulse width mínimo", "Pulse width máximo", "Pulse width médio",
            "Duração da irradiação", "Tamanho do ponto focal",
            "Dose glandular média (evento)", "Dose glandular acumulada", "Exposição na entrada",
            "Camada de semi-atenuação", "Espessura de compressão", "Distância fonte-ponto ref",
            "Área do campo colimado", "Altura do campo", "Largura do campo",
            "Material do anodo", "Filtro principal", "Filtro secundário", "Filtro terciário",
            "Tipo de grade", "Ângulo do posicionador", "Fonte da informação de dose", "UID do evento"
        ]

        # Formatação cabeçalhos
        header_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Adiciona cabeçalhos
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Define larguras das colunas
        column_widths = [
            15, 25, 10, 18, 10, 18, 15, 20, 18, 20, 12, 15, 18,
            10, 12, 12, 12,  # kVp fields
            12, 12, 12, 12,  # mA fields
            15, 12, 15,      # exposure, pulses
            15, 15, 15, 15,  # pulse width fields
            15, 15,          # duration, focal spot
            18, 18, 15, 15, 15, 20,  # dose and geometry
            15, 12, 12,      # field dimensions
            15, 18, 18, 18,  # materials and filters
            15, 15, 25, 35   # grid, angle, dose source, UID
        ]

        for i, width in enumerate(column_widths, 1):
            if i <= len(column_widths):
                if i <= 26:
                    col_letter = chr(64 + i)
                else:
                    col_letter = f"A{chr(64 + i - 26)}"
                ws.column_dimensions[col_letter].width = width

        # Lista de colunas que contêm valores numéricos (índices começando em 1)
        numeric_columns = {
            5,   # Idade
            14,  # kVp
            15,  # kVp mínimo
            16,  # kVp máximo
            17,  # kVp médio
            18,  # Corrente do tubo (mA)
            19,  # mA mínimo
            20,  # mA máximo
            21,  # mA médio
            22,  # Tempo de exposição
            23,  # Número de pulsos
            24,  # Total pulsos registrados
            25,  # Largura do pulso
            26,  # Pulse width mínimo
            27,  # Pulse width máximo
            28,  # Pulse width médio
            29,  # Duração da irradiação
            30,  # Tamanho do ponto focal
            31,  # Dose glandular média (evento)
            32,  # Dose glandular acumulada
            33,  # Exposição na entrada
            34,  # Camada de semi-atenuação
            35,  # Espessura de compressão
            36,  # Distância fonte-ponto ref
            37,  # Área do campo colimado
            38,  # Altura do campo
            39,  # Largura do campo
            43,  # Ângulo do posicionador
        }

        # Processa arquivos DICOM
        row_idx = 2
        processed_count = 0
        error_count = 0

        for i, dicom_file in enumerate(dicom_files, 1):
            try:
                print(f"📄 Processando {i}/{len(dicom_files)}: {os.path.relpath(dicom_file, root_path)}")

                excel_rows = self.extract_excel_data(dicom_file)

                if excel_rows:
                    for excel_row in excel_rows:
                        # Insere dados na planilha
                        for col_idx, value in enumerate(excel_row, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.border = border

                            # Formatação especial para valores numéricos
                            if col_idx in numeric_columns and value is not None and value != '-':
                                if isinstance(value, (int, float)):
                                    # Para números inteiros (como idade, pulsos), não usar decimais
                                    if col_idx in [5, 23, 24]:  # Idade, Número de pulsos, Total pulsos
                                        cell.number_format = '0'
                                    else:
                                        cell.number_format = '0.000'

                        row_idx += 1
                    processed_count += 1

                    if not debug_mode:
                        # Mostra info básica
                        patient_info = f"Patient: {excel_rows[0][0]}" if excel_rows[0][0] != '-' else "No Patient ID"
                        print(f"  ✓ {patient_info}, {len(excel_rows)} eventos de irradiação")
                else:
                    error_count += 1
                    print(f"  ❌ Falha na extração")

            except Exception as e:
                error_count += 1
                print(f"  ❌ Erro: {str(e)}")

        # Salva Excel
        try:
            wb.save(output_file)

            print(f"\n{'=' * 80}")
            print(f"✅ EXCEL DE MAMOGRAFIA GERADO COM SUCESSO!")
            print(f"{'=' * 80}")
            print(f"Arquivo: {output_file}")
            print(f"Arquivos processados: {processed_count}/{len(dicom_files)}")
            print(f"Erros: {error_count}")
            print(f"Total de eventos: {row_idx - 2}")
            print(f"📊 VALORES NUMÉRICOS: Salvos como números (sem unidades) para análise")
            print(f"   • Doses, exposições, ângulos, tempos, etc.")
            print(f"   • Formatação automática com 3 casas decimais")
            print(f"   • Células vazias para valores não encontrados")
            print(f"{'=' * 80}")

            return True

        except Exception as e:
            print(f"❌ Erro ao salvar Excel: {str(e)}")
            return False


def main():
    """Função principal"""
    parser = argparse.ArgumentParser(
        description='DICOMMamographyExtractor - Extração direta de DICOM SR de Mamografia para Excel (VALORES NUMÉRICOS)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos de uso:

1. Processar pasta atual:
   python DICOMMamographyExtractor.py

2. Processar pasta específica:
   python DICOMMamographyExtractor.py --folder /caminho/para/pastas/com/dicoms

3. Com debug ativado:
   python DICOMMamographyExtractor.py --debug

4. Especificar arquivo Excel:
   python DICOMMamographyExtractor.py --output relatorio_mamografia_2024.xlsx

MODIFICAÇÕES NESTA VERSÃO:
✅ Valores numéricos são salvos como números puros (sem unidades)
✅ Doses, exposições, ângulos e medidas ficam como números no Excel
✅ Formatação automática com 3 casas decimais para análise
✅ Células vazias quando valores não são encontrados (ao invés de '-')

O script navega recursivamente pelas pastas, encontra DICOMs SR de dose 
de mamografia e gera diretamente a planilha Excel otimizada para análise numérica.

Dados extraídos incluem:
- Informações do paciente e exame
- Parâmetros técnicos (kVp, mA, tempo de exposição, etc.) - COMO NÚMEROS
- Dose glandular média por evento e acumulada - COMO NÚMEROS  
- Geometria (espessura de compressão, campo colimado) - COMO NÚMEROS
- Equipamento (fabricante, modelo, materiais)
- Lateralidade e projeções (CC, MLO, etc.)
        """
    )

    parser.add_argument('--folder', '-f', default='.',
                        help='Pasta raiz para busca recursiva (padrão: pasta atual)')
    parser.add_argument('--output', '-o', default='mammography_dose_report_numeric.xlsx',
                        help='Nome do arquivo Excel (padrão: mammography_dose_report_numeric.xlsx)')
    parser.add_argument('--debug', '-d', action='store_true',
                        help='Ativa modo debug com informações detalhadas')

    args = parser.parse_args()

    print("=" * 80)
    print("🏥 DICOM MAMMOGRAPHY EXTRACTOR - Extração Direta (VALORES NUMÉRICOS)")
    print("=" * 80)
    print(f"📂 Pasta raiz: {os.path.abspath(args.folder)}")
    print(f"📄 Arquivo Excel: {args.output}")
    print(f"🔍 Debug: {'Ativado' if args.debug else 'Desativado'}")
    print(f"📊 Valores numéricos salvos como números (sem unidades)")
    print("=" * 80)

    if not os.path.exists(args.folder):
        print(f"❌ Pasta não encontrada: {args.folder}")
        return

    # Cria extrator e processa
    extractor = DICOMMamographyExtractor()
    success = extractor.generate_excel_direct(args.folder, args.output, args.debug)

    if success:
        print(f"\n🎯 Processamento de mamografia concluído com sucesso!")
        print(f"📊 Dados específicos extraídos:")
        print(f"   • Dose glandular média por evento e acumulada (NÚMEROS)")
        print(f"   • Lateralidade (Left/Right breast)")
        print(f"   • Projeções (cranio-caudal, MLO)")
        print(f"   • Parâmetros técnicos específicos de mamografia (NÚMEROS)")
        print(f"   • Geometria de compressão e colimação (NÚMEROS)")
        print(f"   • Formatação numérica otimizada para análise")
    else:
        print(f"\n❌ Falha no processamento de mamografia")


if __name__ == "__main__":
    main()
