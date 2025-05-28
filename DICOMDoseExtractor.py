"""
DICOMDoseExtractor.py - Extração direta de DICOM SR para Excel

Este script navega recursivamente por estruturas de pastas, encontra arquivos DICOM SR
de dose de radiação e gera diretamente a planilha Excel sem JSON intermediário.
Extrai apenas os campos necessários para otimizar performance.
"""

import pydicom
import os
import argparse
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class DICOMDirectExcelExtractor:
    """Extrator direto de DICOM SR para Excel"""

    def __init__(self):
        # Códigos DICOM essenciais (apenas os que usamos)
        self.concept_codes = {
            # Dados de irradiação
            'total_dlp': '113813',

            # Aquisição CT
            'ct_acquisition': '113819',
            'acquisition_protocol': '125203',
            'acquisition_type': '113820',
            'comment': '121106',

            # Parâmetros da fonte de raios-X
            'xray_source_params': '113831',
            'kvp': '113733',
            'tube_current': '113734',

            # Dados de dose
            'ct_dose': '113829',
            'mean_ctdivol': '113830',
            'phantom_type': '113835',
            'dlp': '113838',
            'ssde': '113930'
        }

    def find_dicom_files_recursive(self, root_path: str, debug_mode: bool = False) -> list:
        """Busca recursivamente por arquivos DICOM SR"""
        dicom_files = []

        if debug_mode:
            print(f"🔍 Buscando arquivos DICOM em: {root_path}")

        try:
            for root, dirs, files in os.walk(root_path):
                if root == root_path:
                    continue

                for file in files:
                    file_path = os.path.join(root, file)

                    if self.is_dicom_sr_file(file_path):
                        dicom_files.append(file_path)
                        if debug_mode:
                            print(f"  ✓ DICOM encontrado: {file_path}")

        except Exception as e:
            if debug_mode:
                print(f"❌ Erro na busca: {str(e)}")

        return dicom_files

    def is_dicom_sr_file(self, file_path: str) -> bool:
        """Verifica se é um DICOM SR válido rapidamente"""
        try:
            if not os.path.isfile(file_path) or os.path.getsize(file_path) < 132:
                return False

            # Verifica prefixo DICM
            with open(file_path, 'rb') as f:
                f.seek(128)
                if f.read(4) != b'DICM':
                    return False

            # Leitura mínima para verificar se é SR
            ds = pydicom.dcmread(file_path, stop_before_pixels=True, force=True)
            return (hasattr(ds, 'Modality') and ds.Modality == 'SR' and
                    hasattr(ds, 'ContentSequence'))

        except:
            return False

    def calculate_age(self, birth_date_str: str, exam_date_str: str):
        """Calcula idade do paciente"""
        if not birth_date_str or not exam_date_str:
            return '-'

        try:
            birth_date = None
            exam_date = None

            # Formatos de data
            date_formats = ['%b %d, %Y', '%B %d, %Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']
            exam_formats = date_formats + ['%b %d, %Y, %I:%M:%S %p', '%B %d, %Y, %I:%M:%S %p']

            # Parse data nascimento
            for fmt in date_formats:
                try:
                    birth_date = datetime.strptime(birth_date_str.strip(), fmt)
                    break
                except ValueError:
                    continue

            # Parse data exame
            for fmt in exam_formats:
                try:
                    exam_date = datetime.strptime(exam_date_str.strip(), fmt)
                    break
                except ValueError:
                    continue

            # Fallback: extrai apenas anos
            if not birth_date:
                birth_match = re.search(r'(\d{4})', birth_date_str)
                if birth_match:
                    birth_date = datetime(int(birth_match.group(1)), 1, 1)

            if not exam_date:
                exam_match = re.search(r'(\d{4})', exam_date_str)
                if exam_match:
                    exam_date = datetime(int(exam_match.group(1)), 6, 15)

            if birth_date and exam_date:
                age = exam_date.year - birth_date.year
                if (exam_date.month, exam_date.day) < (birth_date.month, birth_date.day):
                    age -= 1
                return age

        except:
            pass

        return '-'

    def format_date(self, date_str: str) -> str:
        """Formata data DICOM"""
        if not date_str or len(date_str) < 8:
            return ""

        try:
            year = date_str[:4]
            month = date_str[4:6]
            day = date_str[6:8]

            months = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            month_name = months[int(month)]
            return f"{month_name} {int(day)}, {year}"
        except:
            return date_str

    def find_content_by_code(self, content_sequence, code_value: str):
        """Encontra item por código DICOM"""
        for item in content_sequence:
            try:
                if (hasattr(item, 'ConceptNameCodeSequence') and
                        item.ConceptNameCodeSequence):
                    concept_code = getattr(item.ConceptNameCodeSequence[0], 'CodeValue', '')
                    if concept_code == code_value:
                        return item
            except:
                continue
        return None

    def get_text_value(self, content_item) -> str:
        """Extrai valor de texto"""
        return getattr(content_item, 'TextValue', '')

    def get_code_meaning(self, content_item) -> str:
        """Extrai code meaning"""
        try:
            if hasattr(content_item, 'ConceptCodeSequence') and content_item.ConceptCodeSequence:
                return getattr(content_item.ConceptCodeSequence[0], 'CodeMeaning', '')
        except:
            pass
        return ""

    def get_numeric_value_with_unit(self, content_item) -> str:
        """Extrai valor numérico com unidade"""
        try:
            if hasattr(content_item, 'MeasuredValueSequence') and content_item.MeasuredValueSequence:
                measured_value = content_item.MeasuredValueSequence[0]
                numeric_value = getattr(measured_value, 'NumericValue', '')

                unit = ''
                if (hasattr(measured_value, 'MeasurementUnitsCodeSequence') and
                        measured_value.MeasurementUnitsCodeSequence):
                    unit_seq = measured_value.MeasurementUnitsCodeSequence[0]
                    unit = getattr(unit_seq, 'CodeMeaning', '')

                if numeric_value and unit:
                    return f"{numeric_value} {unit}"
                elif numeric_value:
                    return str(numeric_value)
        except:
            pass
        return ""

    def extract_excel_data(self, dicom_path: str) -> list:
        """Extrai apenas os dados necessários para o Excel"""
        try:
            ds = pydicom.dcmread(dicom_path)

            if (not hasattr(ds, 'Modality') or ds.Modality != 'SR' or
                    not hasattr(ds, 'ContentSequence')):
                return []

            # Dados básicos do paciente
            patient_id = str(getattr(ds, 'PatientID', ''))
            patient_name = str(getattr(ds, 'PatientName', '')).replace('^', ' ').strip()
            sex = str(getattr(ds, 'PatientSex', ''))

            # Datas
            birth_date_raw = str(getattr(ds, 'PatientBirthDate', ''))
            birth_date = self.format_date(birth_date_raw) if birth_date_raw else ''

            study_date_raw = str(getattr(ds, 'StudyDate', ''))
            study_time_raw = str(getattr(ds, 'StudyTime', ''))
            study_date = self.format_date(study_date_raw) if study_date_raw else ''
            if study_date and study_time_raw and len(study_time_raw) >= 6:
                hour = study_time_raw[:2]
                minute = study_time_raw[2:4]
                second = study_time_raw[4:6]
                study_date = f"{study_date}, {hour}:{minute}:{second}"

            # Calcula idade
            age = self.calculate_age(birth_date, study_date)
            age_value = int(age) if isinstance(age, int) or (isinstance(age, str) and age.isdigit()) else age

            # Patient ID como número se possível
            patient_id_value = int(patient_id) if patient_id and patient_id.isdigit() else (
                patient_id if patient_id else '-')

            # DLP total
            total_dlp = ''
            main_content = ds.ContentSequence

            # Procura DLP total
            for item in main_content:
                try:
                    if (hasattr(item, 'ConceptNameCodeSequence') and
                            item.ConceptNameCodeSequence and
                            getattr(item.ConceptNameCodeSequence[0], 'CodeValue', '') == '113811'):

                        if hasattr(item, 'ContentSequence'):
                            dlp_item = self.find_content_by_code(item.ContentSequence, self.concept_codes['total_dlp'])
                            if dlp_item:
                                total_dlp = self.get_numeric_value_with_unit(dlp_item)
                        break
                except:
                    continue

            # Extrai aquisições
            excel_rows = []
            acquisitions_found = False

            for item in main_content:
                try:
                    if (hasattr(item, 'ConceptNameCodeSequence') and
                            item.ConceptNameCodeSequence and
                            getattr(item.ConceptNameCodeSequence[0], 'CodeValue', '') == self.concept_codes[
                                'ct_acquisition']):

                        acquisitions_found = True

                        if hasattr(item, 'ContentSequence'):
                            acq_content = item.ContentSequence

                            # Dados da aquisição
                            protocol = ''
                            comment = ''
                            acquisition_type = ''
                            phantom_type = ''
                            ctdivol = ''
                            dlp = ''
                            ssde = ''
                            tube_current = ''
                            kvp = ''

                            # Protocol
                            protocol_item = self.find_content_by_code(acq_content,
                                                                      self.concept_codes['acquisition_protocol'])
                            if protocol_item:
                                protocol = self.get_text_value(protocol_item)

                            # Comment
                            comment_item = self.find_content_by_code(acq_content, self.concept_codes['comment'])
                            if comment_item:
                                comment = self.get_text_value(comment_item)

                            # Acquisition Type
                            type_item = self.find_content_by_code(acq_content, self.concept_codes['acquisition_type'])
                            if type_item:
                                acquisition_type = self.get_code_meaning(type_item)

                            # Procura por sub-containers (dose e xray params)
                            for sub_item in acq_content:
                                try:
                                    if (hasattr(sub_item, 'ConceptNameCodeSequence') and
                                            sub_item.ConceptNameCodeSequence):

                                        code = getattr(sub_item.ConceptNameCodeSequence[0], 'CodeValue', '')

                                        # CT Dose
                                        if code == self.concept_codes['ct_dose'] and hasattr(sub_item,
                                                                                             'ContentSequence'):
                                            dose_content = sub_item.ContentSequence

                                            # CTDIvol
                                            ctdivol_item = self.find_content_by_code(dose_content,
                                                                                     self.concept_codes['mean_ctdivol'])
                                            if ctdivol_item:
                                                ctdivol = self.get_numeric_value_with_unit(ctdivol_item)

                                            # DLP
                                            dlp_item = self.find_content_by_code(dose_content,
                                                                                 self.concept_codes['dlp'])
                                            if dlp_item:
                                                dlp = self.get_numeric_value_with_unit(dlp_item)

                                            # Phantom Type
                                            phantom_item = self.find_content_by_code(dose_content,
                                                                                     self.concept_codes['phantom_type'])
                                            if phantom_item:
                                                phantom_type = self.get_code_meaning(phantom_item)

                                            # SSDE
                                            ssde_item = self.find_content_by_code(dose_content,
                                                                                  self.concept_codes['ssde'])
                                            if ssde_item:
                                                ssde = self.get_numeric_value_with_unit(ssde_item)

                                        # X-Ray Source Params (dentro de acquisition params)
                                        elif hasattr(sub_item, 'ContentSequence'):
                                            for param_item in sub_item.ContentSequence:
                                                if (hasattr(param_item, 'ConceptNameCodeSequence') and
                                                        param_item.ConceptNameCodeSequence and
                                                        getattr(param_item.ConceptNameCodeSequence[0], 'CodeValue',
                                                                '') ==
                                                        self.concept_codes['xray_source_params'] and
                                                        hasattr(param_item, 'ContentSequence')):

                                                    xray_content = param_item.ContentSequence

                                                    # Tube Current
                                                    current_item = self.find_content_by_code(xray_content,
                                                                                             self.concept_codes[
                                                                                                 'tube_current'])
                                                    if current_item:
                                                        tube_current = self.get_numeric_value_with_unit(current_item)

                                                    # kVp
                                                    kvp_item = self.find_content_by_code(xray_content,
                                                                                         self.concept_codes['kvp'])
                                                    if kvp_item:
                                                        kvp = self.get_numeric_value_with_unit(kvp_item)

                                                    break
                                except:
                                    continue

                            # Tratamento de valores vazios
                            def safe_value(val):
                                return val if val else '-'

                            # Tratamento especial para comment
                            comment_value = comment if comment and comment.strip() and comment != 'null' else '-'

                            # Cria linha para Excel
                            excel_row = [
                                patient_id_value,  # ID do paciente
                                patient_name or '-',  # Nome do paciente
                                sex or '-',  # Sexo
                                birth_date or '-',  # Data de nascimento
                                age_value,  # Idade
                                protocol or '-',  # Pesquisa de interesse
                                study_date or '-',  # Data do exame
                                comment_value,  # Descrição da série
                                acquisition_type or '-',  # Scan mode
                                tube_current or '-',  # mAs
                                kvp or '-',  # kV
                                ctdivol or '-',  # CTDIvol
                                dlp or '-',  # DLP
                                total_dlp or '-',  # DLP total
                                phantom_type or '-',  # Phantom type
                                ssde or '-',  # SSDE
                                '-'  # Avg scan size (não disponível)
                            ]

                            excel_rows.append(excel_row)

                except:
                    continue

            # Se não encontrou aquisições, cria linha básica
            if not acquisitions_found:
                excel_row = [
                    patient_id_value, patient_name or '-', sex or '-', birth_date or '-', age_value,
                    '-', study_date or '-', '-', '-', '-', '-', '-', '-', total_dlp or '-', '-', '-', '-'
                ]
                excel_rows.append(excel_row)

            return excel_rows

        except Exception as e:
            return []

    def generate_excel_direct(self, root_path: str, output_file: str, debug_mode: bool = False) -> bool:
        """Gera Excel diretamente dos DICOMs"""

        print(f"🔍 Buscando arquivos DICOM em: {os.path.abspath(root_path)}")

        # Busca arquivos DICOM
        dicom_files = self.find_dicom_files_recursive(root_path, debug_mode)

        if not dicom_files:
            print("❌ Nenhum arquivo DICOM SR encontrado")
            return False

        print(f"📊 Encontrados {len(dicom_files)} arquivos DICOM")
        print(f"📄 Gerando Excel: {output_file}")

        # Cria planilha Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatórios DICOM CT"

        # Cabeçalhos
        headers = [
            "ID do paciente", "Nome do paciente", "Sexo", "Data de nascimento", "Idade", "Pesquisa de interesse",
            "Data do exame", "Descrição da série", "Scan mode", "mAs",
            "kV", "CTDIvol", "DLP", "DLP total", "Phantom type", "SSDE", "Avg scan size"
        ]

        # Formatação cabeçalhos
        header_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Adiciona cabeçalhos
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Define larguras das colunas
        column_widths = [15, 25, 10, 18, 10, 20, 18, 20, 15, 10, 10, 10, 10, 10, 15, 10, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Processa arquivos DICOM
        row_idx = 2
        processed_count = 0
        error_count = 0

        for i, dicom_file in enumerate(dicom_files, 1):
            try:
                print(f"📄 Processando {i}/{len(dicom_files)}: {os.path.relpath(dicom_file, root_path)}")

                excel_rows = self.extract_excel_data(dicom_file)

                if excel_rows:
                    for excel_row in excel_rows:
                        # Insere dados na planilha
                        for col_idx, value in enumerate(excel_row, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.border = border
                        row_idx += 1
                    processed_count += 1

                    if not debug_mode:
                        # Mostra info básica
                        patient_info = f"Patient: {excel_rows[0][0]}" if excel_rows[0][0] != '-' else "No Patient ID"
                        print(f"  ✓ {patient_info}, {len(excel_rows)} aquisições")
                else:
                    error_count += 1
                    print(f"  ❌ Falha na extração")

            except Exception as e:
                error_count += 1
                print(f"  ❌ Erro: {str(e)}")

        # Salva Excel
        try:
            wb.save(output_file)

            print(f"\n{'=' * 80}")
            print(f"✅ EXCEL GERADO COM SUCESSO!")
            print(f"{'=' * 80}")
            print(f"Arquivo: {output_file}")
            print(f"Arquivos processados: {processed_count}/{len(dicom_files)}")
            print(f"Erros: {error_count}")
            print(f"Total de linhas: {row_idx - 2}")
            print(f"{'=' * 80}")

            return True

        except Exception as e:
            print(f"❌ Erro ao salvar Excel: {str(e)}")
            return False


def main():
    """Função principal"""
    parser = argparse.ArgumentParser(
        description='DICOMDirectToExcel - Extração direta de DICOM SR para Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos de uso:

1. Processar pasta atual:
   python DICOMDoseExtractor.py

2. Processar pasta específica:
   python DICOMDoseExtractor.py --folder /caminho/para/pastas/com/dicoms

3. Com debug ativado:
   python DICOMDoseExtractor.py --debug

4. Especificar arquivo Excel:
   python DICOMDoseExtractor.py --output relatorio_doses_2024.xlsx

O script navega recursivamente pelas pastas, encontra DICOMs SR de dose 
e gera diretamente a planilha Excel sem JSON intermediário.
        """
    )

    parser.add_argument('--folder', '-f', default='.',
                        help='Pasta raiz para busca recursiva (padrão: pasta atual)')
    parser.add_argument('--output', '-o', default='ct_dose_direct_report.xlsx',
                        help='Nome do arquivo Excel (padrão: ct_dose_direct_report.xlsx)')
    parser.add_argument('--debug', '-d', action='store_true',
                        help='Ativa modo debug com informações detalhadas')

    args = parser.parse_args()

    print("=" * 80)
    print("🏥 DICOM DIRECT TO EXCEL - Extração Direta")
    print("=" * 80)
    print(f"📂 Pasta raiz: {os.path.abspath(args.folder)}")
    print(f"📄 Arquivo Excel: {args.output}")
    print(f"🔍 Debug: {'Ativado' if args.debug else 'Desativado'}")
    print("=" * 80)

    if not os.path.exists(args.folder):
        print(f"❌ Pasta não encontrada: {args.folder}")
        return

    # Cria extrator e processa
    extractor = DICOMDirectExcelExtractor()
    success = extractor.generate_excel_direct(args.folder, args.output, args.debug)

    if success:
        print(f"\n🎯 Processamento concluído com sucesso!")
    else:
        print(f"\n❌ Falha no processamento")


if __name__ == "__main__":
    main()
