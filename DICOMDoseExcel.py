"""
DICOMDoseExcel.py - Conversor de dados de relatórios DICOM de CT para Excel

Este script lê os arquivos JSON gerados pelo DICOMDoseExtractor e cria
uma planilha Excel com as informações organizadas em colunas.
Versão adaptada para JSON consolidado do extrator recursivo.
"""

import json
import os
import glob
import argparse
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def calculate_age(birth_date_str, exam_date_str):
    """Calcula a idade do paciente na época do exame."""
    if not birth_date_str or not exam_date_str:
        return '-'

    try:
        # Tenta parsing com diferentes formatos de data
        birth_date = None
        exam_date = None

        # Parsing da data de nascimento - formatos comuns
        birth_formats = [
            '%b %d, %Y',
            '%B %d, %Y',
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%m/%d/%Y',
        ]

        for fmt in birth_formats:
            try:
                birth_date = datetime.strptime(birth_date_str.strip(), fmt)
                break
            except ValueError:
                continue

        if not birth_date:
            # Se não conseguiu fazer parse, tenta extrair apenas o ano (fallback)
            birth_year_match = re.search(r'(\d{4})', birth_date_str)
            if birth_year_match:
                birth_year = int(birth_year_match.group(1))
                birth_date = datetime(birth_year, 1, 1)  # 1º de janeiro como aproximação
            else:
                return '-'

        # Parsing da data do exame - formatos comuns
        exam_formats = [
            '%b %d, %Y, %I:%M:%S %p',  # "May 5, 2025, 1:20:41 PM"
            '%B %d, %Y, %I:%M:%S %p',  # "May 5, 2025, 1:20:41 PM"
            '%b %d, %Y',  # "May 5, 2025"
            '%B %d, %Y',  # "May 5, 2025"
            '%Y-%m-%d',  # "2025-05-05"
            '%d/%m/%Y',  # "05/05/2025"
            '%m/%d/%Y',  # "05/05/2025"
        ]

        for fmt in exam_formats:
            try:
                exam_date = datetime.strptime(exam_date_str.strip(), fmt)
                break
            except ValueError:
                continue

        if not exam_date:
            # Se não conseguiu fazer parse, tenta extrair apenas o ano (fallback)
            exam_year_match = re.search(r'(\d{4})', exam_date_str)
            if exam_year_match:
                exam_year = int(exam_year_match.group(1))
                exam_date = datetime(exam_year, 6, 15)  # Meio do ano como aproximação
            else:
                return '-'

        age = exam_date.year - birth_date.year

        # Verifica se o aniversário já passou na data do exame
        if (exam_date.month, exam_date.day) < (birth_date.month, birth_date.day):
            age -= 1  # Subtrai 1 se o aniversário ainda não chegou

        return str(age)

    except Exception as e:
        # Em caso de erro, tenta o cálculo simples por ano
        birth_year_match = re.search(r'(\d{4})', birth_date_str)
        exam_year_match = re.search(r'(\d{4})', exam_date_str)

        if birth_year_match and exam_year_match:
            birth_year = int(birth_year_match.group(1))
            exam_year = int(exam_year_match.group(1))
            age = exam_year - birth_year
            return str(age)

        return '-'


def extract_scan_info(acquisition):
    """Extrai informações de aquisição para cada linha, exatamente como estão no JSON"""
    scan_info = {}

    # Protocolo (será usado como Pesquisa de interesse)
    scan_info['protocol'] = acquisition.get('protocol', '-')

    # Descrição da série (APENAS comment, sem fallback)
    # Tratamento especial para comment (garantir que null se torne '-')
    comment = acquisition.get('comment')
    # Múltiplas verificações para garantir que qualquer valor "vazio" se torne '-'
    if comment is None or comment == '' or (isinstance(comment, str) and comment.strip() == '') or comment == 'null':
        scan_info['description'] = '-'
    else:
        scan_info['description'] = comment

    # Scan mode (tipo de aquisição)
    scan_info['scan_mode'] = acquisition.get('acquisition_type', '-')

    # Phantom type
    ct_dose = acquisition.get('ct_dose', {}) or {}  # Usa {} se for None
    scan_info['phantom_type'] = ct_dose.get('phantom_type', '-')

    # CTDI vol - valor exato como está no JSON
    ctdivol = ct_dose.get('mean_ctdivol')
    scan_info['ctdivol'] = ctdivol if ctdivol is not None else '-'

    # DLP - valor exato como está no JSON
    dlp = ct_dose.get('dlp')
    scan_info['dlp'] = dlp if dlp is not None else '-'

    # SSDE - valor exato como está no JSON, com tratamento especial para None
    ssde = ct_dose.get('size_specific_dose')
    scan_info['ssde'] = ssde if ssde is not None else '-'

    # Dados da fonte de raios X
    xray_params = acquisition.get('xray_source_params', {}) or {}  # Usa {} se for None

    # Tube current - valor exato como está no JSON
    tube_current = xray_params.get('tube_current')
    scan_info['tube_current'] = tube_current if tube_current is not None else '-'

    # kV - valor exato como está no JSON
    kv = xray_params.get('kvp')
    scan_info['kv'] = kv if kv is not None else '-'

    # Avg scan size - não está disponível geralmente
    scan_info['avg_scan_size'] = '-'

    return scan_info


def json_to_excel(json_file, output_file="ct_dose_dicom_report.xlsx"):
    """Converte os dados JSON de DICOM para Excel"""

    # Lê o arquivo JSON (pode ser consolidado ou legado)
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Verifica se é o formato consolidado novo (com metadata e reports)
        if isinstance(data, dict) and 'metadata' in data and 'reports' in data:
            reports = data['reports']
            print(f"✓ JSON consolidado detectado com {len(reports)} relatórios")
            print(f"  Gerado em: {data['metadata'].get('generated_at', 'N/A')}")
        elif isinstance(data, list):
            # Formato legado (lista direta de relatórios)
            reports = data
            print(f"✓ JSON legado detectado com {len(reports)} relatórios")
        else:
            # Objeto único
            reports = [data]
            print(f"✓ Relatório único detectado")

        print(f"✓ Processando {len(reports)} relatórios do arquivo '{os.path.basename(json_file)}'")

    except Exception as e:
        print(f"❌ Erro ao ler JSON: {str(e)}")
        return False

    # Cria uma nova planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatórios DICOM CT"

    # Define os cabeçalhos
    headers = [
        "ID do paciente", "Nome do paciente", "Sexo", "Data de nascimento", "Idade", "Pesquisa de interesse",
        "Data do exame", "Descrição da série", "Scan mode", "mAs",
        "kV", "CTDIvol", "DLP", "DLP total", "Phantom type", "SSDE", "Avg scan size"
    ]

    # Formata os cabeçalhos
    header_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")  # Azul claro para DICOM
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Adiciona cabeçalhos na primeira linha
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Define larguras das colunas
    ws.column_dimensions['A'].width = 15  # ID do paciente
    ws.column_dimensions['B'].width = 25  # Nome do paciente
    ws.column_dimensions['C'].width = 10  # Sexo
    ws.column_dimensions['D'].width = 18  # Data de nascimento
    ws.column_dimensions['E'].width = 10  # Idade
    ws.column_dimensions['F'].width = 20  # Pesquisa de interesse
    ws.column_dimensions['G'].width = 18  # Data do exame
    ws.column_dimensions['H'].width = 20  # Descrição da série
    ws.column_dimensions['I'].width = 15  # Scan mode
    ws.column_dimensions['J'].width = 10  # mAs
    ws.column_dimensions['K'].width = 10  # kV
    ws.column_dimensions['L'].width = 10  # CTDIvol
    ws.column_dimensions['M'].width = 10  # DLP
    ws.column_dimensions['N'].width = 10  # DLP total
    ws.column_dimensions['O'].width = 15  # Phantom type
    ws.column_dimensions['P'].width = 10  # SSDE
    ws.column_dimensions['Q'].width = 15  # Avg scan size

    # Linha atual para inserção
    row_idx = 2

    # Processa cada relatório
    for report in reports:
        essential = report.get('essential', {})

        # Obtém informações básicas do paciente/estudo
        patient_id = essential.get('patient_id', '')
        patient_name = essential.get('patient_name', '')
        sex = essential.get('sex', '')
        birth_date = essential.get('birth_date', '')
        study_date = essential.get('study_date', '')

        # DLP total - direto do JSON, sem extração
        irradiation = report.get('irradiation', {})
        total_dlp = irradiation.get('total_dlp', '')

        # Processa cada aquisição/série como uma linha
        acquisitions = report.get('acquisitions', [])
        if acquisitions:
            for acquisition in acquisitions:
                # Obtém informações desta aquisição específica
                scan_info = extract_scan_info(acquisition)

                # Insere valores na planilha com tratamento explícito para None/null
                patient_id_value = int(patient_id) if patient_id and patient_id.isdigit() else (patient_id if patient_id else '-')
                ws.cell(row=row_idx, column=1, value=patient_id_value)
                ws.cell(row=row_idx, column=2, value=patient_name if patient_name is not None else '-')
                ws.cell(row=row_idx, column=3, value=sex if sex is not None else '-')
                ws.cell(row=row_idx, column=4, value=birth_date if birth_date is not None else '-')

                # Calcula a idade com base na data de nascimento e data do exame
                age = calculate_age(birth_date, study_date)
                age_value = int(age) if age != '-' and age.isdigit() else age
                ws.cell(row=row_idx, column=5, value=age_value)

                ws.cell(row=row_idx, column=6, value=scan_info['protocol'])
                ws.cell(row=row_idx, column=7, value=study_date if study_date is not None else '-')
                # Descrição da série - tratamento especial para garantir '-' em caso de null
                description_value = scan_info['description']
                # Verificação extra rigorosa para garantir que não seja null, string vazia, espaços, etc.
                is_empty = (description_value is None or description_value == '' or
                            description_value.strip() == '' or description_value == 'null')
                ws.cell(row=row_idx, column=8, value='-' if is_empty else description_value)
                ws.cell(row=row_idx, column=9, value=scan_info['scan_mode'])
                ws.cell(row=row_idx, column=10, value=scan_info['tube_current'])
                ws.cell(row=row_idx, column=11, value=scan_info['kv'])
                ws.cell(row=row_idx, column=12, value=scan_info['ctdivol'])
                ws.cell(row=row_idx, column=13, value=scan_info['dlp'])
                ws.cell(row=row_idx, column=14, value=total_dlp if total_dlp is not None else '-')
                ws.cell(row=row_idx, column=15, value=scan_info['phantom_type'])
                ws.cell(row=row_idx, column=16, value=scan_info['ssde'])
                ws.cell(row=row_idx, column=17, value=scan_info['avg_scan_size'])

                # Aplica borda a todas as células
                for col_idx in range(1, 18):
                    ws.cell(row=row_idx, column=col_idx).border = border

                row_idx += 1
        else:
            # Se não houver aquisições, adiciona pelo menos uma linha com dados básicos
            patient_id_value = int(patient_id) if patient_id and patient_id.isdigit() else (patient_id if patient_id else '-')
            ws.cell(row=row_idx, column=1, value=patient_id_value)
            ws.cell(row=row_idx, column=2, value=patient_name if patient_name is not None else '-')
            ws.cell(row=row_idx, column=3, value=sex if sex is not None else '-')
            ws.cell(row=row_idx, column=4, value=birth_date if birth_date is not None else '-')

            # Calcula a idade para esta linha também
            age = calculate_age(birth_date, study_date)
            age_value = int(age) if age != '-' and age.isdigit() else age
            ws.cell(row=row_idx, column=5, value=age_value)

            ws.cell(row=row_idx, column=6, value='-')
            ws.cell(row=row_idx, column=7, value=study_date if study_date is not None else '-')
            ws.cell(row=row_idx, column=8, value='-')
            ws.cell(row=row_idx, column=9, value='-')
            ws.cell(row=row_idx, column=10, value='-')
            ws.cell(row=row_idx, column=11, value='-')
            ws.cell(row=row_idx, column=12, value='-')
            ws.cell(row=row_idx, column=13, value='-')
            ws.cell(row=row_idx, column=14, value=total_dlp if total_dlp is not None else '-')
            ws.cell(row=row_idx, column=15, value='-')
            ws.cell(row=row_idx, column=16, value='-')
            ws.cell(row=row_idx, column=17, value='-')

            # Aplica borda a todas as células
            for col_idx in range(1, 18):
                ws.cell(row=row_idx, column=col_idx).border = border

            row_idx += 1

    # Salva a planilha
    try:
        wb.save(output_file)
        print(f"✅ Planilha Excel DICOM salva com sucesso: '{output_file}'")
        print(f"   Total de linhas geradas: {row_idx - 2}")
        return True
    except Exception as e:
        print(f"❌ Erro ao salvar planilha Excel: {str(e)}")
        return False


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='Conversor de relatórios JSON de DICOM CT para Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos de uso:

1. Arquivo JSON consolidado (novo formato):
   python DICOMDoseExcel.py dicom_reports_consolidated_20241228_143022.json

2. Especificar arquivo de saída:
   python DICOMDoseExcel.py dados.json --output relatorio_dose_2024.xlsx

3. Arquivo JSON legado (formato antigo):
   python DICOMDoseExcel.py ct_reports_dicom_all.json

O script detecta automaticamente se o JSON é consolidado (novo) ou legado.
        """
    )

    parser.add_argument('json_file',
                        help='Arquivo JSON com dados DICOM (consolidado ou legado)')
    parser.add_argument('--output', type=str, default='ct_dose_dicom_report.xlsx',
                        help='Nome do arquivo Excel de saída (padrão: ct_dose_dicom_report.xlsx)')

    args = parser.parse_args()

    print(f"\n{'=' * 80}")
    print(f"DICOMDoseExcel - Conversor de Relatórios JSON DICOM para Excel")
    print(f"{'=' * 80}")
    print(f"Arquivo JSON: {args.json_file}")
    print(f"Arquivo Excel: {args.output}")
    print(f"{'=' * 80}\n")

    # Verifica se arquivo existe
    if not os.path.exists(args.json_file):
        print(f"❌ Arquivo JSON não encontrado: {args.json_file}")
    else:
        json_to_excel(args.json_file, args.output)
